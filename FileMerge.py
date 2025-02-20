import datetime
import pandas as pd
import numpy as np
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
from openpyxl import load_workbook
from openpyxl.styles import Font, Color

# Create a Tkinter root window
root = Tk()
# Hide the root window
root.withdraw()

try:
    # Open the file picker dialog
    clockIn_File = askopenfilename()
    payRoll_File = askopenfilename()

    # Check if a file path was selected
    if not clockIn_File or not payRoll_File:
        print("No file selected.")
        raise SystemExit

    # Read the Excel files
    df1 = pd.read_excel(clockIn_File)
    df2 = pd.read_excel(payRoll_File)

    # Keep only 'Employee name', 'Ticket Date' and 'Agency' columns in df2
    df2 = df2[['Employee Name', 'Ticket Date', 'Agency', 'Supervisors Name', 'PM Assigned', 'JobNo|Customer|Description']]

    # Convert 'Ticket Date' to datetime in both dataframes
    df1['Ticket Date'] = pd.to_datetime(df1['Ticket Date'])
    df2['Ticket Date'] = pd.to_datetime(df2['Ticket Date'])

    # If 'Actual Clock In Time' and 'Actual Clock Out Time' are not datetime, convert them
    df1['Actual Clock In Time'] = pd.to_datetime(df1['Actual Clock In Time'])
    df1['Actual Clock Out Time'] = pd.to_datetime(df1['Actual Clock Out Time'])

    # Set empty 'Actual Clock Out Time' to 'Actual Clock In Time'
    df1.loc[df1['Actual Clock Out Time'].isnull(), 'Actual Clock Out Time'] = df1['Actual Clock In Time']

    # Calculate 'Actual Hours Worked' as the difference between 'Actual Clock Out Time' and 'Actual Clock In Time',
    # converted to hours
    df1['Actual Hours Worked'] = (df1['Actual Clock Out Time'] - df1['Actual Clock In Time']).dt.total_seconds() / 3600

    # Round 'Actual Hours Worked' to 2 decimal places
    df1['Actual Hours Worked'] = df1['Actual Hours Worked'].apply(lambda x: round(x, 2))

    # # Round 'Actual Hours Worked' using the 7-minute rule
    # df1['Actual Hours Worked'] = df1['Actual Hours Worked'].apply(lambda x: round((x * 60 + 7) // 15 / 4, 2))

    # Add 'Day of the Week' column
    df1['Day of the Week'] = df1['Ticket Date'].dt.day_name()

    # Merge dataframes based on 'Employee name' and 'Ticket Date'
    merged_df = pd.merge(df1, df2, on=['Employee Name', 'Ticket Date'], how='left')

    # If 'Agency' is blank, fill with 'CSI'
    merged_df['Agency'] = merged_df['Agency'].fillna('CSI')

    # Create a new dataframe for rows with errors
    errors_df = merged_df[(merged_df['Actual Clock In Time'].isna()) |
                          (merged_df['Actual Clock Out Time'].isna()) |
                          (merged_df['Actual Hours Worked'] < 8)].copy()


    # Create the 'Error Description' column
    def generate_error_desc(row):
        if pd.isnull(row['Actual Clock In Time']) and pd.isnull(row['Actual Clock Out Time']):
            return 'No Clock In or Clock Out Time'
        elif pd.isnull(row['Actual Clock In Time']):
            return 'No Clock In'
        elif pd.isnull(row['Actual Clock Out Time']):
            return 'No Clock Out'
        elif row['Actual Hours Worked'] < 8:
            return 'Less Than 8 Hours'
        else:
            return np.nan


    errors_df['Error Description'] = errors_df.apply(generate_error_desc, axis=1)

    # Convert 'Ticket Date' back to 'mm/dd/yyyy' format
    merged_df['Ticket Date'] = merged_df['Ticket Date'].dt.strftime('%m/%d/%Y')
    errors_df['Ticket Date'] = errors_df['Ticket Date'].dt.strftime('%m/%d/%Y')

    # Write the dataframes into a new Excel file with two sheets
    # C:\\Test\\Payroll.xlsx
    with pd.ExcelWriter('C:\\Users\\tj-fo\\Desktop\\Test\\Results.xlsx') as writer:
        merged_df.to_excel(writer, sheet_name='Payroll', index=False)
        errors_df.to_excel(writer, sheet_name='Errors', index=False)

    # Load workbook
    wb = load_workbook('C:\\Users\\tj-fo\\Desktop\\Test\\Payroll.xlsx')  # Update with the correct file path
    # Read the Excel file
    df = pd.read_excel('C:\\Users\\tj-fo\\Desktop\\Test\\Payroll.xlsx')

    # Select the sheets
    sheet1 = wb['Payroll']

    # Create a red bold font
    red_bold_font = Font(color="FF0000", bold=True)

    # Check each cell in column E (5th column) for both sheets
    for sheet in [sheet1]:
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=5):
            for cell in row:
                if cell.column_letter == 'D' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock In Time?'
                    cell.font = red_bold_font
                elif cell.column_letter == 'E' and (cell.value is None or cell.value == ''):
                    cell.value = 'Clock Out Time?'
                    cell.font = red_bold_font

    # Set column widths
    sheet1.column_dimensions['A'].width = 11.26
    sheet1.column_dimensions['B'].width = 26.14
    sheet1.column_dimensions['C'].width = 31.86
    sheet1.column_dimensions['D'].width = 19
    sheet1.column_dimensions['E'].width = 20.43
    sheet1.column_dimensions['F'].width = 18.71
    sheet1.column_dimensions['G'].width = 20.86
    sheet1.column_dimensions['H'].width = 32.57
    sheet1.column_dimensions['I'].width = 28.71
    sheet1.column_dimensions['J'].width = 22.86
    sheet1.column_dimensions['K'].width = 57.86

    # C:\\Test\\Payroll.xlsx
    # Save workbook
    wb.save('C:\\Users\\tj-fo\\Desktop\\Test\\Payroll.xlsx')
    # Load workbook
    wb = load_workbook('C:\\Users\\tj-fo\\Desktop\\Test\\Payroll.xlsx')  # Update with the correct file path
    # Read the Excel file
    df = pd.read_excel('C:\\Users\\tj-fo\\Desktop\\Test\\Payroll.xlsx')
    # Select the sheet
    sheet2 = wb['Payroll']

    # Create a red bold font
    red_bold_font = Font(color="FF0000", bold=True)

    # Set column widths
    sheet2.column_dimensions['A'].width = 11.26
    sheet2.column_dimensions['B'].width = 26.14
    sheet2.column_dimensions['C'].width = 31.86
    sheet2.column_dimensions['D'].width = 19
    sheet2.column_dimensions['E'].width = 20.43
    sheet2.column_dimensions['F'].width = 18.71
    sheet2.column_dimensions['G'].width = 20.86
    sheet2.column_dimensions['H'].width = 32.57
    sheet2.column_dimensions['I'].width = 28.71
    sheet2.column_dimensions['J'].width = 22.86
    sheet2.column_dimensions['K'].width = 57.86

    def round_time(dt):
        # Calculate the number of minutes past the last 15-minute mark
        minutes = (dt.minute % 15) * 60 + dt.second

        # If the number of minutes is less than 7, round down; otherwise, round up
        if minutes < 7 * 60:
            dt = dt - datetime.timedelta(minutes=dt.minute % 15, seconds=dt.second)
        else:
            dt = dt + datetime.timedelta(minutes=15 - dt.minute % 15, seconds=-dt.second)

        return dt

    # Convert the 'Ticket Date', 'Actual Clock In Time', and 'Actual Clock Out Time' columns to datetime
    df['Ticket Date'] = pd.to_datetime(df['Ticket Date'])
    df['Actual Clock In Time'] = pd.to_datetime(df['Actual Clock In Time'])
    df['Actual Clock Out Time'] = pd.to_datetime(df['Actual Clock Out Time'])

    # Calculate the total hours worked for each job
    df['Total Hours Worked'] = (df['Actual Clock Out Time'] - df['Actual Clock In Time']).dt.total_seconds() / 3600

    # Create a list to hold the results
    results = []

    # Group by 'Employee Name', 'Quote/Job Number Number', 'Agency', and 'Ticket Date'
    grouped = df.groupby(['Employee Name', 'Quote/Job Number Number', 'Agency', df['Ticket Date'].dt.date])

    for name, group in grouped:
        total_hours = group['Total Hours Worked'].sum()
        regular_hours = round(min(8, total_hours), 2)
        overtime_hours = round(max(0, total_hours - 8), 2)
        # Deduct 30 minutes for lunch break if the employee worked for more than 5 hours
        if total_hours > 5:
            total_hours -= 0.5

        # If the Ticket Date is on a Saturday or Sunday, all hours are overtime
        if group['Ticket Date'].dt.dayofweek.iloc[0] >= 5:
            overtime_hours = round(total_hours, 2)
            regular_hours = 0

        results.append(pd.DataFrame({
            'Employee Name': [name[0]],
            'JobNo|Customer|Description': [name[1]],
            'Agency': [name[2]],
            'Ticket Date': [group['Ticket Date'].dt.date.iloc[0]],
            'Day': [group['Ticket Date'].dt.day_name().iloc[0]],
            'Regular Hours': [regular_hours],
            'Overtime Hours': [overtime_hours],
            'Actual Clock In Time': [group['Actual Clock In Time'].iloc[0]],
            'Actual Clock Out Time': [group['Actual Clock Out Time'].iloc[0]]
            # 'Supervisors Name': [supervisors_name]
        }))

    # Concatenate all the results into a single dataframe
    result = pd.concat(results)

    # # Open the save file dialog
    # file_path = asksaveasfilename()

    # # Check if a file path was selected
    # if file_path:
    #     # Open the file and write some content
    #     with open(file_path, 'w') as file:
    #         file.write('This is some example content.')
    #
    #     print(f"File saved at: {file_path}")
    # else:
    #     print("No file selected.")

    # C:\\Test\\Payroll.xlsx
    # Save the result to a new Excel file
    result.to_excel('C:\\Users\\tj-fo\\Desktop\\Test\\Payroll.xlsx', index=False)
    # Destroy the Tkinter root window
    root.destroy()
except Exception as e:
    print("An error occurred:", str(e))
    raise SystemExit
